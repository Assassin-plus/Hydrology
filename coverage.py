# Requirements: Spatial Analyst Extension
import arcpy
from arcpy import env
from arcpy.sa import *
import os
import sys
import numpy

XLSX = True
try:
    from openpyxl import Workbook
except ImportError:
    print("No openpyxl module found. Saving output to txt file instead.")
    XLSX = False

DEBUG = True
TEST = False
SAVE = False

############################################################################################################
# Set environment settings
############################################################################################################
# Check out the ArcGIS Spatial Analyst extension license
arcpy.CheckOutExtension("Spatial")
arcpy.CheckOutExtension("3D")
# Overwrite output
arcpy.env.overwriteOutput = True

# Get current path
path = os.getcwd()
# Set workspace
# env.workspace = path
# print("Workspace set to: " + path)

# Set input and output folders
inputFolder = path + "\\data"
outputFolder = path + "\\output"
print("Input folder set to: " + inputFolder)
print("Output folder set to: " + outputFolder)

# Set overall input
input = inputFolder + "\\north_america_2015_v2\\NA_NALCMS_2015_v2_land_cover_30m\\NA_NALCMS_2015_v2_land_cover_30m.tif"

# Get list of subinput
subinput = []
subinput_dir = inputFolder + "\\Boundary\\Ohio"
for file in os.listdir(subinput_dir):
    if file.endswith(".shp"):
        subinput.append(file)
if DEBUG:
    print("Subinput list: " + str(subinput))
    print("Number of subinput: " + str(len(subinput)))

# Set overall output
# for every subinput, create a folder in output folder
suboutput = []
for i in range(len(subinput)):
    suboutput.append(outputFolder + "\\" + subinput[i][:-4])
    if not os.path.exists(suboutput[i]):
        os.makedirs(suboutput[i])
if DEBUG:
    print("Suboutput list: " + str(suboutput))
    print("Number of suboutput: " + str(len(suboutput)))

if TEST:
    subinput = ["3170000.shp"]
    suboutput = [outputFolder + "\\" + sub[:-4] for sub in subinput]
    print("test subinput: " + str(subinput))
    print("test suboutput: " + str(suboutput))

# data structure for each subbasin
subbasin = {}
# Save output to excel
if XLSX:
    wb = Workbook()
    wb.create_sheet('Sheet1', 0)
    ws = wb.active
    ws.cell(1,1).value = "Subbasin Number"
    ws.cell(1,2).value = "Temperate or sub-polar needleleaf forest Coverage"
    ws.cell(1,3).value = "Sub-polar taiga needleleaf forest Coverage"
    ws.cell(1,4).value = "Tropical or sub-tropical broadleaf evergreen forest Coverage"
    ws.cell(1,5).value = "Tropical or sub-tropical broadleaf deciduous forest Coverage"
    ws.cell(1,6).value = "Temperate or sub-polar broadleaf deciduous forest Coverage"
    ws.cell(1,7).value = "Mixed forest Coverage"
    ws.cell(1,8).value = "Tropical or sub-tropical shrubland Coverage"
    ws.cell(1,9).value = "Temperate or sub-polar shrubland Coverage"
    ws.cell(1,10).value = "Tropical or sub-tropical grassland Coverage"
    ws.cell(1,11).value = "Temperate or sub-polar grassland Coverage"
    ws.cell(1,12).value = "Sub-polar or polar shrubland-lichen-moss Coverage"
    ws.cell(1,13).value = "Sub-polar or polar grassland-lichen-moss Coverage"
    ws.cell(1,14).value = "Sub-polar or polar barren-lichen-moss Coverage"
    ws.cell(1,15).value = "Wetland Coverage"
    ws.cell(1,16).value = "Cropland Coverage"
    ws.cell(1,17).value = "Barren land Coverage"
    ws.cell(1,18).value = "Urban and built-up Coverage"
    ws.cell(1,19).value = "Water Coverage"
    ws.cell(1,20).value = "Permanent snow and ice Coverage"
else:
    f = open(outputFolder + "\\result.txt", "w")
    f.write("Subbasin Number" + "\t" +
            "Temperate or sub-polar needleleaf forest Coverage" + "\t" +
            "Sub-polar taiga needleleaf forest Coverage" + "\t" +
            "Tropical or sub-tropical broadleaf evergreen forest Coverage" + "\t" +
            "Tropical or sub-tropical broadleaf deciduous forest Coverage" + "\t" +
            "Temperate or sub-polar broadleaf deciduous forest Coverage" + "\t" +
            "Mixed forest Coverage" + "\t" +
            "Tropical or sub-tropical shrubland Coverage" + "\t" +
            "Temperate or sub-polar shrubland Coverage" + "\t" +
            "Tropical or sub-tropical grassland Coverage" + "\t" +
            "Temperate or sub-polar grassland Coverage" + "\t" +
            "Sub-polar or polar shrubland-lichen-moss Coverage" + "\t" +
            "Sub-polar or polar grassland-lichen-moss Coverage" + "\t" +
            "Sub-polar or polar barren-lichen-moss Coverage" + "\t" +
            "Wetland Coverage" + "\t" +
            "Cropland Coverage" + "\t" +
            "Barren land Coverage" + "\t" +
            "Urban and built-up Coverage" + "\t" +
            "Water Coverage" + "\t" +
            "Permanent snow and ice Coverage" + "\n")

###########################################################################################################
# Part1: Hydrological Analysis
############################################################################################################
# Step0: Iterate through subinput
############################################################################################################
# data structure for each subbasin
result = [[] for i in range(len(subinput))]

for i in range(len(subinput)):
    try:
        result[i] = [0 for j in range(20)]
        print("Processing subbasin: " + subinput[i])
        # set workspace to suboutput
        env.workspace = suboutput[i]
        print("Workspace set to: " + suboutput[i])
    ############################################################################################################
    # Step1: Extract by mask
    ############################################################################################################
        # Set local variables
        inRaster = input
        inMaskData = inputFolder + "\\Boundary\\Ohio\\" + subinput[i]
        if DEBUG:
            print("inRaster: " + inRaster)
            print("inMaskData: " + inMaskData)
        # Execute ExtractByMask
        outExtractByMask = ExtractByMask(inRaster, inMaskData)
        if SAVE:
            outExtractByMask.save(suboutput[i] + "\\dem.tif")
            if DEBUG:
                print("ExtractByMask saved to: " + suboutput[i] + "\\dem.tif")
    ############################################################################################################
    # Step2: subbasin coverage
    ############################################################################################################
        # Set local variables
        inZoneData = outExtractByMask
        # Execute GetCount
        result[i][0] = arcpy.GetCount_management(inZoneData)
        print("Number of cells: " + str(result[i][0].getOutput(0)))
        # convert inZoneData to numpy array
        arr = arcpy.RasterToNumPyArray(inZoneData)
        # get unique values and their frequency of occurrence
        for j in range(1,19):
            result[i][j] = numpy.count_nonzero(arr == j)
            if DEBUG:
                print("Value " + str(j) + ": " + str(result[i][j]))
        result[i][0] = sum(result[i][1:19])
        if DEBUG:
            print("Total: " + str(result[i][0]))
        # get percentage
        for j in range(1,19):
            result[i][j] = float(result[i][j]) / float(result[i][0])
            if DEBUG:
                print("Value " + str(j) + ": " + str(result[i][j]))
        
    
    ############################################################################################################
    # Finish, save result
    ############################################################################################################
        if XLSX:
            ws.cell(i+2,1).value = subinput[i][:-4]
            for j in range(1,19):
                ws.cell(i+2,j+1).value = result[i][j]
        else:
            f.write(subinput[i][:-4] + "\t" +
                    str(result[i][1]) + "\t" +
                    str(result[i][2]) + "\t" +
                    str(result[i][3]) + "\t" +
                    str(result[i][4]) + "\t" +
                    str(result[i][5]) + "\t" +
                    str(result[i][6]) + "\t" +
                    str(result[i][7]) + "\t" +
                    str(result[i][8]) + "\t" +
                    str(result[i][9]) + "\t" +
                    str(result[i][10]) + "\t" +
                    str(result[i][11]) + "\t" +
                    str(result[i][12]) + "\t" +
                    str(result[i][13]) + "\t" +
                    str(result[i][14]) + "\t" +
                    str(result[i][15]) + "\t" +
                    str(result[i][16]) + "\t" +
                    str(result[i][17]) + "\t" +
                    str(result[i][18]) + "\t" +
                    str(result[i][19]) + "\n")
        
        

        ############################################################################################################
        # Move the file to done folder
        ############################################################################################################
        # os.mkdir(path + "\\done\\" + subinput[i][:-4])
        # move all files in subinput[i] to done folder
        # os.rename(subinput_dir + "\\" + subinput[i], path + "\\done\\" + subinput[i])
    except:
        print("Error processing subbasin: " + subinput[i])
        #print details of error
        import traceback
        tb = sys.exc_info()[2]
        tbinfo = traceback.format_tb(tb)[0]
        pymsg = "ERRORS:\nTraceback Info:\n" + tbinfo + "\nError Info:\n    " + str(sys.exc_info()[1])
        print(pymsg)

    
if XLSX:
    wb.save(outputFolder + "\\result_coverage.xlsx")
    print("Output saved to: " + outputFolder + "\\result_coverage.xlsx")
else:
    f.close()
    print("Output saved to: " + outputFolder + "\\result_coverage.txt")

print("All done!")