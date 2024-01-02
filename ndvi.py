# Requirements: Spatial Analyst Extension
import arcpy
from arcpy import env
from arcpy.sa import *
import os
import sys
import numpy

XLSX = True
try:
    from openpyxl import Workbook
except ImportError:
    print("No openpyxl module found. Saving output to txt file instead.")
    XLSX = False

DEBUG = True
TEST = True
SAVE = False

############################################################################################################
# Set environment settings
############################################################################################################
# Check out the ArcGIS Spatial Analyst extension license
arcpy.CheckOutExtension("Spatial")
arcpy.CheckOutExtension("3D")
# Overwrite output
arcpy.env.overwriteOutput = True

# Get current path
path = os.getcwd()
# Set workspace
# env.workspace = path
# print("Workspace set to: " + path)

# Set input and output folders
inputFolder = path + "\\data"
outputFolder = path + "\\output"
print("Input folder set to: " + inputFolder)
print("Output folder set to: " + outputFolder)

# Set overall input
input = inputFolder + "\\global_ndvi\\"
subinputyear = []
for file in os.listdir(input):
    if file.endswith(".tif"):
        subinputyear.append(file)
        

# Get list of subinput
subinput = []
subinput_dir = inputFolder + "\\Boundary\\Ohio"
for file in os.listdir(subinput_dir):
    if file.endswith(".shp"):
        subinput.append(file)
if DEBUG:
    print("Subinput list: " + str(subinput))
    print("Number of subinput: " + str(len(subinput)))

# Set overall output
# for every subinput, create a folder in output folder
suboutput = []
for i in range(len(subinput)):
    suboutput.append(outputFolder + "\\" + subinput[i][:-4])
    if not os.path.exists(suboutput[i]):
        os.makedirs(suboutput[i])
if DEBUG:
    print("Suboutput list: " + str(suboutput))
    print("Number of suboutput: " + str(len(suboutput)))

if TEST:
    subinput = ["3078000.shp"]
    suboutput = [outputFolder + "\\" + sub[:-4] for sub in subinput]
    print("test subinput: " + str(subinput))
    print("test suboutput: " + str(suboutput))

# data structure for each subbasin
subbasin = {}
# Save output to excel
if XLSX:
    wb = Workbook()
    wb.create_sheet('Sheet1', 0)
    ws = wb.active
    ws.cell(1,1).value = "Subbasin Number"
    ws.cell(1,2).value = "Average NDVI"
else:
    f = open(outputFolder + "\\result.txt", "w")
    

###########################################################################################################
# Part1: Hydrological Analysis
############################################################################################################
# Step0: Iterate through subinput
############################################################################################################
# data structure for each subbasin
result = [[] for i in range(len(subinput))]
# parallel processing
import multiprocessing
from multiprocessing import Pool

def process_subbasin(i):
    try:
        print("Processing subbasin: " + subinput[i])
        # set workspace to suboutput
        env.workspace = suboutput[i]
        print("Workspace set to: " + suboutput[i])
    ############################################################################################################
    # Step1: Extract by mask
    ############################################################################################################
        # Set local variables
        for j in subinputyear:
            inRaster = inputFolder + "\\global_ndvi\\" + j
            inMaskData = inputFolder + "\\Boundary\\Ohio\\" + subinput[i]
            if DEBUG:
                print("inRaster: " + inRaster)
                print("inMaskData: " + inMaskData)
            # Execute ExtractByMask
            outExtractByMask = ExtractByMask(inRaster, inMaskData)
            if outExtractByMask is None:
                print("Error extracting by mask")
            else:
                result[i].append(float(arcpy.GetRasterProperties_management(outExtractByMask, "MEAN").getOutput(0)))
            if DEBUG:
                print("Mean: " + str(result[i][-1]))

        subbasin_result = sum(result[i])/float(len(result[i]))/10000.0
        print("Average NDVI: " + str(subbasin_result))
    
    ############################################################################################################
    # Finish, save result
    ############################################################################################################
        if XLSX:
            ws.cell(i+2,1).value = subinput[i][:-4]
            ws.cell(i+2,2).value = subbasin_result
        else:
            f.write(subinput[i][:-4] + "\t")
            f.write(str(subbasin_result) + "\t")
            f.write("\n")
            
        ############################################################################################################
        # Move the file to done folder
        ############################################################################################################
        # os.mkdir(path + "\\done\\" + subinput[i][:-4])
        # move all files in subinput[i] to done folder
        # os.rename(subinput_dir + "\\" + subinput[i], path + "\\done\\" + subinput[i])
    except:
        print("Error processing subbasin: " + subinput[i])
        #print details of error
        import traceback
        tb = sys.exc_info()[2]
        tbinfo = traceback.format_tb(tb)[0]
        pymsg = "ERRORS:\nTraceback Info:\n" + tbinfo + "\nError Info:\n    " + str(sys.exc_info()[1])
        print(pymsg)

def save():
    if XLSX:
        wb.save(outputFolder + "\\result_ndvi.xlsx")
        print("Output saved to: " + outputFolder + "\\result_ndvi.xlsx")
    else:
        f.close()
        print("Output saved to: " + outputFolder + "\\result_ndvi.txt")

    print("All done!")

if __name__ == '__main__':
    """ pool = Pool()
    pool.map(process_subbasin, range(len(subinput)))
    pool.close()
    pool.join() """
    for i in range(len(subinput)):
        process_subbasin(i)
    save()
