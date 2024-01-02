# Requirements: Spatial Analyst Extension
import arcpy
from arcpy import env
from arcpy.sa import *
import os
import sys

XLSX = True
try:
    from openpyxl import Workbook
except ImportError:
    print("No openpyxl module found. Saving output to txt file instead.")
    XLSX = False

DEBUG = True
TEST = False
SAVE = False

############################################################################################################
# Set environment settings
############################################################################################################
# Check out the ArcGIS Spatial Analyst extension license
arcpy.CheckOutExtension("Spatial")
arcpy.CheckOutExtension("3D")
# Overwrite output
arcpy.env.overwriteOutput = True

# Get current path
path = os.getcwd()
# Set workspace
# env.workspace = path
# print("Workspace set to: " + path)

# Set input and output folders
inputFolder = path + "\\data"
outputFolder = path + "\\output"
print("Input folder set to: " + inputFolder)
print("Output folder set to: " + outputFolder)

# Set overall input
input = inputFolder + "\\Source\\ohio_90.tif"

# Get list of subinput
subinput = []
subinput_dir = inputFolder + "\\Boundary\\Ohio"
for file in os.listdir(subinput_dir):
    if file.endswith(".shp"):
        subinput.append(file)
if DEBUG:
    print("Subinput list: " + str(subinput))
    print("Number of subinput: " + str(len(subinput)))

# Set overall output
# for every subinput, create a folder in output folder
suboutput = []
for i in range(len(subinput)):
    suboutput.append(outputFolder + "\\" + subinput[i][:-4])
    if not os.path.exists(suboutput[i]):
        os.makedirs(suboutput[i])
if DEBUG:
    print("Suboutput list: " + str(suboutput))
    print("Number of suboutput: " + str(len(suboutput)))

if TEST:
    subinput = ["15243_USGS_03203600_Boundary_Line.shp", "3170000.shp"]
    suboutput = [outputFolder + "\\" + sub[:-4] for sub in subinput]
    print("test subinput: " + str(subinput))
    print("test suboutput: " + str(suboutput))

# data structure for each subbasin
subbasin = {}
# Save output to excel
if XLSX:
    wb = Workbook()
    wb.create_sheet('Sheet1', 0)
    ws = wb.active
    ws.cell(1,1).value = 'Subbasin Number'
    ws.cell(1,2).value = 'Subbasin Average Height'
    ws.cell(1,3).value = 'Subbasin Area(km^2)'
    ws.cell(1,4).value = 'Subbasin River Length(km)'
    ws.cell(1,5).value = 'Subbasin River Density(km/km^2)'
    ws.cell(1,6).value = 'Subbasin River Average Slope(Degree)'
    ws.cell(1,7).value = 'Subbasin River Average TWI'
else:
    f = open(outputFolder + "\\result.txt", "w")
    f.write("Subbasin Number\tSubbasin Average Height\tSubbasin Area(km^2)\tSubbasin River Length(km)\tSubbasin River Density(km/km^2)\tSubbasin River Average Slope(Degree)\tSubbasin River Average TWI\n")

############################################################################################################
# Part1: Hydrological Analysis
############################################################################################################
# Step0: Iterate through subinput
############################################################################################################
# data structure for each subbasin
result = [[] for i in range(len(subinput))]

for i in range(len(subinput)):
    try:
        print("Processing subbasin: " + subinput[i])
        # set workspace to suboutput
        env.workspace = suboutput[i]
        print("Workspace set to: " + suboutput[i])
    ############################################################################################################
    # Step1: Extract by mask
    ############################################################################################################
        # Set local variables
        inRaster = input
        inMaskData = inputFolder + "\\Boundary\\Ohio\\" + subinput[i]
        if DEBUG:
            print("inRaster: " + inRaster)
            print("inMaskData: " + inMaskData)
        # Execute ExtractByMask
        outExtractByMask = ExtractByMask(inRaster, inMaskData)
        if SAVE:
            outExtractByMask.save(suboutput[i] + "\\dem.tif")
            if DEBUG:
                print("ExtractByMask saved to: " + suboutput[i] + "\\dem.tif")
    ############################################################################################################
    # Step2: Fill
    ############################################################################################################
        outFill = Fill(outExtractByMask)
        if SAVE:
            outFill.save(suboutput[i] + "\\fill.tif")
            if DEBUG:
                print("Fill saved to: " + suboutput[i] + "\\fill.tif")
    ############################################################################################################
    # Step3: Flow Direction
    ############################################################################################################
        outFlowDirection = FlowDirection(outFill, "NORMAL")
        if SAVE:
            outFlowDirection.save(suboutput[i] + "\\dir.tif")
            if DEBUG:
                print("FlowDirection saved to: " + suboutput[i] + "\\dir.tif")
    ############################################################################################################
    # Step4: Flow Accumulation
    ############################################################################################################
        outFlowAccumulation = FlowAccumulation(outFlowDirection)
        if SAVE:
            outFlowAccumulation.save(suboutput[i] + "\\acc.tif")
            if DEBUG:
                print("FlowAccumulation saved to: " + suboutput[i] + "\\acc.tif")
    ############################################################################################################
    # Step5: AccFlow
    ############################################################################################################
        outAccFlow = Con(outFlowAccumulation > 1000, 1)
        if SAVE:
            outAccFlow.save(suboutput[i] + "\\accflow.tif")
            if DEBUG:
                print("AccFlow saved to: " + suboutput[i] + "\\accflow.tif")
    ############################################################################################################
    # Step6: Stream Link
    ############################################################################################################
        outStreamLink = StreamLink(outAccFlow, outFlowDirection)
        if SAVE:
            outStreamLink.save(suboutput[i] + "\\link.tif")
            if DEBUG:
                print("StreamLink saved to: " + suboutput[i] + "\\link.tif")
    ############################################################################################################
    # Step7: Stream to Feature
    ############################################################################################################
        outStreamToFeature = StreamToFeature(outStreamLink, outFlowDirection, suboutput[i] + "\\stream.shp", "SIMPLIFY")
        if DEBUG:
            print("StreamToFeature saved to: " + suboutput[i] + "\\stream.shp")
    ############################################################################################################
    # Step8: Watershed
    ############################################################################################################
        ## TODO May there be a problem with value logcurve.
        outWatershed = Watershed(outFlowDirection, outStreamLink)
        if SAVE:
            outWatershed.save(suboutput[i] + "\\watershed.tif")
            if DEBUG:
                print("Watershed saved to: " + suboutput[i] + "\\watershed.tif")
    ############################################################################################################
    # Step9: Raster to Polygon
    ############################################################################################################
        outRasterToPolygon = arcpy.conversion.RasterToPolygon(outWatershed, suboutput[i] + "\\watershed.shp", "NO_SIMPLIFY", "VALUE")
        if DEBUG:
            print("RasterToPolygon saved to: " + suboutput[i] + "\\watershed.shp")
    ############################################################################################################
    # Step10: Raster Domain
    ############################################################################################################
        try:
            outRasterDomain = arcpy.ddd.RasterDomain(outWatershed, suboutput[i] + "\\boundary.shp", "LINE")
            if DEBUG:
                print("RasterDomain saved to: " + suboutput[i] + "\\boundary.shp")
        except:
            print("RasterDomain failed. Most likely because the license is not available.")
        

    ############################################################################################################
    # Part2: Terrain Analysis
    ############################################################################################################
    # Step1: Average Height of Subbasin
    ############################################################################################################
        meanHeight = arcpy.GetRasterProperties_management(outExtractByMask, "MEAN")
        result[i].append(meanHeight)
        if DEBUG:
            print("Average Height of Subbasin: " + str(meanHeight))
    ############################################################################################################
    # Step2: Area of Subbasin
    ############################################################################################################
        # deprecated method after ArcGIS 10.6
        # area = arcpy.management.CalculateGeometryAttributes(outRasterToPolygon, "AREA")
        arcpy.AddField_management(outRasterToPolygon,'AREA','DOUBLE')
        arcpy.CalculateField_management(outRasterToPolygon,'AREA','!shape.area!','PYTHON')
        #area = arcpy.ListFields(outRasterToPolygon, "AREA")
        #print(area[0].Value)
        areas = arcpy.SearchCursor(outRasterToPolygon, fields="AREA")
        area_sum = sum([area.getValue("AREA") for area in areas])/1000000
        result[i].append(area_sum)
        if DEBUG:
            print("Area of Subbasin: " + str(area_sum))
    ############################################################################################################
    # Step3: River Length of Subbasin
    ############################################################################################################
        arcpy.AddField_management(outStreamToFeature,'LENGTH','DOUBLE')
        arcpy.CalculateField_management(outStreamToFeature,'LENGTH','!shape.length!','PYTHON')
        length = arcpy.SearchCursor(outStreamToFeature, fields="LENGTH")
        riverLength = sum([length.getValue("LENGTH") for length in length])/1000
        result[i].append(riverLength)
        if DEBUG:
            print("River Length of Subbasin: " + str(riverLength))
    ############################################################################################################
    # Step4: River Density of Subbasin
    ############################################################################################################
        riverDensity = float(riverLength) / float(area_sum)
        result[i].append(riverDensity)
        if DEBUG:
            print("River Density of Subbasin: " + str(riverDensity))
    ############################################################################################################
    # Step5: River Average Slope of Subbasin
    ############################################################################################################
        outSlope = arcpy.ddd.Slope(outExtractByMask, "DEGREE")
        riverSlope = arcpy.GetRasterProperties_management(outSlope, "MEAN")
        result[i].append(riverSlope)
        if DEBUG:
            print("River Average Slope of Subbasin: " + str(riverSlope))
    ############################################################################################################
    # Step6: Average Slope of Subbasin After Fill
    ############################################################################################################
        outFillSlope = Slope(outFill, "DEGREE")
        if SAVE:
            outFillSlope.save(suboutput[i] + "\\slope_fill.tif")
            if DEBUG:
                print("FillSlope saved to: " + suboutput[i] + "\\slope_fill.tif")
        fillSlope = arcpy.GetRasterProperties_management(outFillSlope, "MEAN")
        # result[i].append(fillSlope)
        if DEBUG:
            print("Average Slope of Subbasin After Fill: " + str(fillSlope))
    ############################################################################################################
    # Step7: River Average TWI of Subbasin
    ############################################################################################################
        Dir = outFlowDirection
        Acc = outFlowAccumulation
        outSCA = Con(Acc == 0,1,Acc) * 90*90 / Con( Dir == 1,90,Con(Dir == 4,90,Con(Dir == 16,90,Con(Dir == 64,90,Con(Dir == 2,90* SquareRoot(2),Con(Dir == 8,90* SquareRoot(2),Con(Dir == 32,90* SquareRoot(2),Con(Dir == 128,90* SquareRoot(2)))))))))
        """ c = Con("dir.tif" == 128,90* SquareRoot(2))
        d = Con("dir.tif" == 32,90* SquareRoot(2),c)
        e = Con("dir.tif" == 8,90* SquareRoot(2),d)
        f = Con("dir.tif" == 2,90* SquareRoot(2),e)
        g = Con("dir.tif" == 64,90,f)
        h = Con("dir.tif" == 16,90,g)
        i = Con("dir.tif" == 4,90,h)
        b = Con( "dir.tif" == 1,90,i)
        a = Con("acc.tif" == 0,1,"acc.tif")
        outSCA = a * 90 * 90 / b  """
        if SAVE:
            outSCA.save(suboutput[i] + "\\SCA.tif")
            if DEBUG:
                print("SCA saved to: " + suboutput[i] + "\\SCA.tif")
        outTWI = Ln(outSCA/Tan(Con(outFillSlope<=0, 0.00001, Con(outFillSlope>0,outFillSlope*3.1415926/180))))
        if SAVE:
            outTWI.save(suboutput[i] + "\\TWI.tif")
            if DEBUG:
                print("TWI saved to: " + suboutput[i] + "\\TWI.tif")
        riverTWI = arcpy.GetRasterProperties_management(outTWI, "MEAN")
        result[i].append(riverTWI)
        if DEBUG:
            print("River Average TWI of Subbasin: " + str(riverTWI))

    ############################################################################################################
    # Finish, save result
    ############################################################################################################
        if XLSX:
            ws.cell(i+2,1).value = subinput[i][:-4]
            ws.cell(i+2,2).value = result[i][0].getOutput(0)
            ws.cell(i+2,3).value = result[i][1]
            ws.cell(i+2,4).value = result[i][2]
            ws.cell(i+2,5).value = result[i][3]
            ws.cell(i+2,6).value = result[i][4].getOutput(0)
            ws.cell(i+2,7).value = result[i][5].getOutput(0)
            print("ws.cell(" + str(i+2) + ",1).value = " + subinput[i][:-4])
            print("ws.cell(" + str(i+2) + ",2).value = " + str(result[i][0].getOutput(0)))
            print("ws.cell(" + str(i+2) + ",3).value = " + str(result[i][1]))
            print("ws.cell(" + str(i+2) + ",4).value = " + str(result[i][2]))
            print("ws.cell(" + str(i+2) + ",5).value = " + str(result[i][3]))
            print("ws.cell(" + str(i+2) + ",6).value = " + str(result[i][4].getOutput(0)))
            print("ws.cell(" + str(i+2) + ",7).value = " + str(result[i][5].getOutput(0)))
        else:
            f.write(subinput[i][:-4] + "\t" + 
                    str(result[i][0].getOutput(0)) + "\t" + 
                    str(result[i][1]) + "\t" + 
                    str(result[i][2]) + "\t" + 
                    str(result[i][3]) + "\t" + 
                    str(result[i][4].getOutput(0)) + "\t" + 
                    str(result[i][5].getOutput(0)) + "\n")
        
        ############################################################################################################
        # Clean up
        ############################################################################################################
        #while True:
        try:
            arcpy.Delete_management(outExtractByMask)
            arcpy.Delete_management(outFill)
            arcpy.Delete_management(outFlowDirection)
            arcpy.Delete_management(outFlowAccumulation)
            arcpy.Delete_management(outAccFlow)
            arcpy.Delete_management(outStreamLink)
            arcpy.Delete_management(outStreamToFeature)
            arcpy.Delete_management(outWatershed)
            arcpy.Delete_management(outRasterToPolygon)
            arcpy.Delete_management(outRasterDomain)
            arcpy.Delete_management(outSlope)
            arcpy.Delete_management(outFillSlope)
            arcpy.Delete_management(outSCA)
            arcpy.Delete_management(outTWI)
        except:
            print("Error deleting temporary files.")

        ############################################################################################################
        # Move the file to done folder
        ############################################################################################################
        # os.mkdir(path + "\\done\\" + subinput[i][:-4])
        # move all files in subinput[i] to done folder
        os.rename(subinput_dir + "\\" + subinput[i], path + "\\done\\" + subinput[i])
    except:
        print("Error processing subbasin: " + subinput[i])
        #print details of error
        import traceback
        tb = sys.exc_info()[2]
        tbinfo = traceback.format_tb(tb)[0]
        pymsg = "ERRORS:\nTraceback Info:\n" + tbinfo + "\nError Info:\n    " + str(sys.exc_info()[1])
        print(pymsg)

    
if XLSX:
    wb.save(outputFolder + "\\result_water.xlsx")
    print("Output saved to: " + outputFolder + "\\result_water.xlsx")
else:
    f.close()
    print("Output saved to: " + outputFolder + "\\result_water.txt")

print("All done!")