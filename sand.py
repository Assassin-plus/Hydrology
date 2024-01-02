# Requirements: Spatial Analyst Extension
import arcpy
from arcpy import env
from arcpy.sa import *
import os
import sys
import numpy

XLSX = True
try:
    from openpyxl import Workbook
except ImportError:
    print("No openpyxl module found. Saving output to txt file instead.")
    XLSX = False

DEBUG = True
TEST = False
SAVE = False

############################################################################################################
# Set environment settings
############################################################################################################
# Check out the ArcGIS Spatial Analyst extension license
arcpy.CheckOutExtension("Spatial")
arcpy.CheckOutExtension("3D")
# Overwrite output
arcpy.env.overwriteOutput = True

# Get current path
path = os.getcwd()
# Set workspace
# env.workspace = path
# print("Workspace set to: " + path)

# Set input and output folders
inputFolder = path + "\\data"
outputFolder = path + "\\output"
print("Input folder set to: " + inputFolder)
print("Output folder set to: " + outputFolder)

# Set overall input
input = inputFolder + "\\sand"
subinputlayer_dir = []
for file in os.listdir(input):
    if file.endswith(".tif"):
        subinputlayer_dir.append(file)
if DEBUG:
    print("Subinput layer list: " + str(subinputlayer_dir))
    print("Number of subinput layer: " + str(len(subinputlayer_dir)))

# Get list of subinput
subinput = []
subinput_dir = inputFolder + "\\Boundary\\Ohio"
for file in os.listdir(subinput_dir):
    if file.endswith(".shp"):
        subinput.append(file)
if DEBUG:
    print("Subinput list: " + str(subinput))
    print("Number of subinput: " + str(len(subinput)))

# Set overall output
# for every subinput, create a folder in output folder
suboutput = []
for i in range(len(subinput)):
    suboutput.append(outputFolder + "\\" + subinput[i][:-4])
    if not os.path.exists(suboutput[i]):
        os.makedirs(suboutput[i])
if DEBUG:
    print("Suboutput list: " + str(suboutput))
    print("Number of suboutput: " + str(len(suboutput)))

if TEST:
    subinput = ["3170000.shp"]
    suboutput = [outputFolder + "\\" + sub[:-4] for sub in subinput]
    print("test subinput: " + str(subinput))
    print("test suboutput: " + str(suboutput))

# data structure for each subbasin
subbasin = {}
# Save output to excel
if XLSX:
    wb = Workbook()
    wb.create_sheet('Sheet1', 0)
    ws = wb.active
    ws.cell(1,1).value = "Subbasin Number"
    ws.cell(1,2).value = "CLAY1"
    ws.cell(1,3).value = "CLAY2"
    ws.cell(1,4).value = "CLAY3"
    ws.cell(1,5).value = "CLAY4"
    ws.cell(1,6).value = "CLAY5"
    ws.cell(1,7).value = "CLAY6"
    ws.cell(1,8).value = "GRAV1"
    ws.cell(1,9).value = "GRAV2"
    ws.cell(1,10).value = "GRAV3"
    ws.cell(1,11).value = "GRAV4"
    ws.cell(1,12).value = "GRAV5"
    ws.cell(1,13).value = "GRAV6"
    ws.cell(1,14).value = "SILT1"
    ws.cell(1,15).value = "SILT2"
    ws.cell(1,16).value = "SILT3"
    ws.cell(1,17).value = "SILT4"
    ws.cell(1,18).value = "SILT5"
    ws.cell(1,19).value = "SILT6"
    ws.cell(1,20).value = "SAND1"
    ws.cell(1,21).value = "SAND2"
    ws.cell(1,22).value = "SAND3"
    ws.cell(1,23).value = "SAND4"
    ws.cell(1,24).value = "SAND5"
    ws.cell(1,25).value = "SAND6"
else:
    f = open(outputFolder + "\\result.txt", "w")
    f.write("Subbasin Number\t" +
            "CLAY1\t" +
            "CLAY2\t" +
            "CLAY3\t" +
            "CLAY4\t" +
            "CLAY5\t" +
            "CLAY6\t" +
            "GRAV1\t" +
            "GRAV2\t" +
            "GRAV3\t" +
            "GRAV4\t" +
            "GRAV5\t" +
            "GRAV6\t" +
            "SILT1\t" +
            "SILT2\t" +
            "SILT3\t" +
            "SILT4\t" +
            "SILT5\t" +
            "SILT6\t" +
            "SAND1\t" +
            "SAND2\t" +
            "SAND3\t" +
            "SAND4\t" +
            "SAND5\t" +
            "SAND6\t" )
    f.write("\n")

###########################################################################################################
# Part1: Hydrological Analysis
############################################################################################################
# Step0: Iterate through subinput
############################################################################################################
# data structure for each subbasin
result = [[] for i in range(len(subinput))]

for i in range(len(subinput)):
    try:
        result[i] = [0 for j in range(24)]
        print("Processing subbasin: " + subinput[i])
        # set workspace to suboutput
        env.workspace = suboutput[i]
        print("Workspace set to: " + suboutput[i])
    ############################################################################################################
    # Step1: Extract by mask
    ############################################################################################################
        # for each subinput, extract all subinputlayer
        for j in range(len(subinputlayer_dir)):
            # Extract by mask
            inRaster = input + "\\" + subinputlayer_dir[j]
            inMaskData = inputFolder + "\\Boundary\\Ohio\\" + subinput[i]
            outExtractByMask = ExtractByMask(inRaster, inMaskData)
            # Save the output 
            if SAVE:
                outExtractByMask.save(subinputlayer_dir[j][:-4] + "_extract.tif")
                print("Extracted " + subinputlayer_dir[j] + " to " + subinputlayer_dir[j][:-4] + "_extract.tif")

            ############################################################################################################
            # Step2: Calculate mean
            ############################################################################################################
            result[i][j] = arcpy.GetRasterProperties_management(outExtractByMask, "MEAN")
            if DEBUG:
                print("Mean of " + subinputlayer_dir[j] + " is " + str(result[i][j]))
    
    ############################################################################################################
    # Finish, save result
    ############################################################################################################
        if XLSX:
            ws.cell(i+2,1).value = subinput[i][:-4]
            for j in range(0,24):
                ws.cell(i+2,j+2).value = float(result[i][j].getOutput(0)) / 100.0
        else:
            f.write(subinput[i][:-4] + "\t")
            for j in range(24):
                f.write(str(float(result[i][j].getOutput(0)) / 100.0) + "\t")
            f.write("\n")
        

        ############################################################################################################
        # Move the file to done folder
        ############################################################################################################
        # os.mkdir(path + "\\done\\" + subinput[i][:-4])
        # move all files in subinput[i] to done folder
        # os.rename(subinput_dir + "\\" + subinput[i], path + "\\done\\" + subinput[i])
    except:
        print("Error processing subbasin: " + subinput[i])
        #print details of error
        import traceback
        tb = sys.exc_info()[2]
        tbinfo = traceback.format_tb(tb)[0]
        pymsg = "ERRORS:\nTraceback Info:\n" + tbinfo + "\nError Info:\n    " + str(sys.exc_info()[1])
        print(pymsg)

    
if XLSX:
    wb.save(outputFolder + "\\result_sand.xlsx")
    print("Output saved to: " + outputFolder + "\\result_sand.xlsx")
else:
    f.close()
    print("Output saved to: " + outputFolder + "\\result_sand.txt")

print("All done!")